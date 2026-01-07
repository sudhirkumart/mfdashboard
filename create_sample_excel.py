#!/usr/bin/env python3
"""
Create a sample Excel file for testing import functionality
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    print("Creating sample Excel file for import testing...")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Holdings"

    # Add headers
    headers = [
        'PAN', 'Account Name', 'Scheme Code', 'Scheme Name', 'Folio',
        'Category', 'Units', 'Invested Amount', 'Current NAV',
        'Current Value', 'Gain/Loss', 'Return %', 'NAV Date'
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    # Add sample data
    data = [
        ['AAAAA1111A', 'Test User', '119551', 'HDFC Top 100 Fund - Direct Plan - Growth',
         '12345678', 'Equity - Large Cap', 100.5, 65325.00, 720.50, 72410.25, 7085.25, 10.85, '03-01-2024'],
        ['AAAAA1111A', 'Test User', '118989', 'Axis Bluechip Fund - Direct Plan - Growth',
         '87654321', 'Equity - Large Cap', 250.0, 10625.00, 48.25, 12062.50, 1437.50, 13.53, '03-01-2024'],
        ['BBBBB2222B', 'Another User', '120503', 'SBI Small Cap Fund - Direct Plan - Growth',
         '99887766', 'Equity - Small Cap', 80.0, 4968.00, 95.80, 7664.00, 2696.00, 54.27, '03-01-2024'],
    ]

    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15
    ws.column_dimensions['D'].width = 45  # Scheme name

    # Save file
    output_file = 'data/sample_import.xlsx'
    wb.save(output_file)

    print(f"✓ Created: {output_file}")
    print(f"  - Sheet: Holdings")
    print(f"  - Rows: {len(data) + 1} (including header)")
    print(f"  - Accounts: 2 PANs")
    print(f"  - Holdings: {len(data)}")
    print("\nYou can now:")
    print("  1. Open http://localhost:8000")
    print("  2. Click 'Import' button")
    print("  3. Select data/sample_import.xlsx")
    print("  4. Watch it load!")

except ImportError:
    print("⚠ openpyxl not installed")
    print("Install with: pip install openpyxl")
    print("\nAlternatively, you can:")
    print("  1. Export CSV from dashboard")
    print("  2. Open in Excel")
    print("  3. Save as .xlsx")
    print("  4. Import the .xlsx file")

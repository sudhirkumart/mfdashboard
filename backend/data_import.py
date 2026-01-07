"""
Data Import Module
Imports portfolio data from CSV, Excel files
"""

import csv
import json
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime
import logging

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class DataImporter:
    """Import portfolio data from various file formats"""

    def __init__(self):
        """Initialize data importer"""
        pass

    def import_holdings_csv(self, csv_path: str) -> List[Dict]:
        """
        Import holdings from CSV file

        Args:
            csv_path: Path to CSV file

        Returns:
            List of holdings dictionaries
        """
        logger.info(f"Importing holdings from CSV: {csv_path}")

        holdings = []

        with open(csv_path, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)

            for row in reader:
                holding = {
                    'scheme_code': row.get('Scheme Code', ''),
                    'scheme_name': row.get('Scheme Name', ''),
                    'folio': row.get('Folio', ''),
                    'category': row.get('Category', ''),
                    'units': float(row.get('Units', 0)),
                    'invested': float(row.get('Invested Amount', 0)),
                    'current_nav': float(row.get('Current NAV', 0)),
                    'current_value': float(row.get('Current Value', 0)),
                    'nav_date': row.get('NAV Date', ''),
                    'transactions': []  # Will be populated from transactions CSV
                }
                holdings.append(holding)

        logger.info(f"Imported {len(holdings)} holdings from CSV")
        return holdings

    def import_transactions_csv(self, csv_path: str) -> List[Dict]:
        """
        Import transactions from CSV file

        Args:
            csv_path: Path to CSV file

        Returns:
            List of transaction dictionaries
        """
        logger.info(f"Importing transactions from CSV: {csv_path}")

        transactions = []

        with open(csv_path, 'r', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)

            for row in reader:
                transaction = {
                    'date': row.get('Date', ''),
                    'scheme_code': row.get('Scheme Code', ''),
                    'scheme_name': row.get('Scheme Name', ''),
                    'folio': row.get('Folio', ''),
                    'type': row.get('Type', '').lower(),
                    'units': float(row.get('Units', 0)),
                    'nav': float(row.get('NAV', 0)),
                    'amount': float(row.get('Amount', 0))
                }
                transactions.append(transaction)

        logger.info(f"Imported {len(transactions)} transactions from CSV")
        return transactions

    def import_holdings_excel(self, excel_path: str, sheet_name: str = "Holdings") -> List[Dict]:
        """
        Import holdings from Excel file

        Args:
            excel_path: Path to Excel file
            sheet_name: Name of the sheet to read

        Returns:
            List of holdings dictionaries
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel import. Install with: pip install openpyxl")

        logger.info(f"Importing holdings from Excel: {excel_path}")

        workbook = openpyxl.load_workbook(excel_path, data_only=True)

        if sheet_name not in workbook.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
            sheet_name = workbook.sheetnames[0]
            logger.info(f"Using sheet: {sheet_name}")

        sheet = workbook[sheet_name]
        holdings = []

        # Get headers from first row
        headers = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                headers[col_idx] = cell.value

        # Read data rows
        for row_idx in range(2, sheet.max_row + 1):
            row_data = {}
            for col_idx, header in headers.items():
                cell_value = sheet.cell(row_idx, col_idx).value
                row_data[header] = cell_value if cell_value is not None else ''

            # Skip empty rows or total rows
            if not row_data.get('Scheme Name') or row_data.get('PAN', '').upper() == 'TOTAL':
                continue

            holding = {
                'scheme_code': str(row_data.get('Scheme Code', '')),
                'scheme_name': str(row_data.get('Scheme Name', '')),
                'folio': str(row_data.get('Folio', '')),
                'category': str(row_data.get('Category', '')),
                'units': float(row_data.get('Units', 0)) if row_data.get('Units') else 0,
                'invested': float(row_data.get('Invested Amount', 0) or row_data.get('Invested', 0)) if row_data.get('Invested Amount') or row_data.get('Invested') else 0,
                'current_nav': float(row_data.get('Current NAV', 0)) if row_data.get('Current NAV') else 0,
                'current_value': float(row_data.get('Current Value', 0)) if row_data.get('Current Value') else 0,
                'nav_date': str(row_data.get('NAV Date', '')),
                'transactions': []
            }
            holdings.append(holding)

        logger.info(f"Imported {len(holdings)} holdings from Excel")
        workbook.close()
        return holdings

    def import_transactions_excel(self, excel_path: str, sheet_name: str = "Transactions") -> List[Dict]:
        """
        Import transactions from Excel file

        Args:
            excel_path: Path to Excel file
            sheet_name: Name of the sheet to read

        Returns:
            List of transaction dictionaries
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel import. Install with: pip install openpyxl")

        logger.info(f"Importing transactions from Excel: {excel_path}")

        workbook = openpyxl.load_workbook(excel_path, data_only=True)

        if sheet_name not in workbook.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
            return []

        sheet = workbook[sheet_name]
        transactions = []

        # Get headers from first row
        headers = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                headers[col_idx] = cell.value

        # Read data rows
        for row_idx in range(2, sheet.max_row + 1):
            row_data = {}
            for col_idx, header in headers.items():
                cell_value = sheet.cell(row_idx, col_idx).value
                row_data[header] = cell_value if cell_value is not None else ''

            # Skip empty rows
            if not row_data.get('Scheme Name'):
                continue

            transaction = {
                'date': str(row_data.get('Date', '')),
                'scheme_code': str(row_data.get('Scheme Code', '')),
                'scheme_name': str(row_data.get('Scheme Name', '')),
                'folio': str(row_data.get('Folio', '')),
                'type': str(row_data.get('Type', '')).lower(),
                'units': float(row_data.get('Units', 0)) if row_data.get('Units') else 0,
                'nav': float(row_data.get('NAV', 0)) if row_data.get('NAV') else 0,
                'amount': float(row_data.get('Amount', 0)) if row_data.get('Amount') else 0
            }
            transactions.append(transaction)

        logger.info(f"Imported {len(transactions)} transactions from Excel")
        workbook.close()
        return transactions

    def csv_to_portfolio(self, holdings_csv: str, transactions_csv: Optional[str] = None,
                        pan: str = "IMPORTED", account_name: str = "Imported Account") -> Dict:
        """
        Convert CSV files to portfolio JSON format

        Args:
            holdings_csv: Path to holdings CSV file
            transactions_csv: Optional path to transactions CSV file
            pan: PAN number for the account
            account_name: Name for the account

        Returns:
            Portfolio data dictionary
        """
        logger.info("Converting CSV to portfolio format")

        holdings = self.import_holdings_csv(holdings_csv)

        transactions = []
        if transactions_csv and Path(transactions_csv).exists():
            transactions = self.import_transactions_csv(transactions_csv)

        # Group transactions by scheme
        transactions_by_scheme = {}
        for txn in transactions:
            scheme_code = txn['scheme_code']
            if scheme_code not in transactions_by_scheme:
                transactions_by_scheme[scheme_code] = []
            transactions_by_scheme[scheme_code].append(txn)

        # Add transactions to holdings
        for holding in holdings:
            scheme_code = holding['scheme_code']
            if scheme_code in transactions_by_scheme:
                holding['transactions'] = transactions_by_scheme[scheme_code]

        # Create portfolio structure
        portfolio = {
            'portfolio_name': f'Imported Portfolio - {datetime.now().strftime("%Y-%m-%d")}',
            'created_date': datetime.now().isoformat(),
            'accounts': {
                pan: {
                    'name': account_name,
                    'email': '',
                    'holdings': holdings
                }
            }
        }

        logger.info(f"Created portfolio with {len(holdings)} holdings")
        return portfolio

    def excel_to_portfolio(self, excel_path: str, pan: str = "IMPORTED",
                          account_name: str = "Imported Account") -> Dict:
        """
        Convert Excel file to portfolio JSON format

        Args:
            excel_path: Path to Excel file
            pan: PAN number for the account
            account_name: Name for the account

        Returns:
            Portfolio data dictionary
        """
        logger.info("Converting Excel to portfolio format")

        holdings = self.import_holdings_excel(excel_path, sheet_name="Holdings")
        transactions = []

        try:
            transactions = self.import_transactions_excel(excel_path, sheet_name="Transactions")
        except Exception as e:
            logger.warning(f"Could not import transactions: {e}")

        # Group transactions by scheme
        transactions_by_scheme = {}
        for txn in transactions:
            scheme_code = txn['scheme_code']
            if scheme_code not in transactions_by_scheme:
                transactions_by_scheme[scheme_code] = []
            transactions_by_scheme[scheme_code].append(txn)

        # Add transactions to holdings
        for holding in holdings:
            scheme_code = holding['scheme_code']
            if scheme_code in transactions_by_scheme:
                holding['transactions'] = transactions_by_scheme[scheme_code]

        # Create portfolio structure
        portfolio = {
            'portfolio_name': f'Imported Portfolio - {datetime.now().strftime("%Y-%m-%d")}',
            'created_date': datetime.now().isoformat(),
            'accounts': {
                pan: {
                    'name': account_name,
                    'email': '',
                    'holdings': holdings
                }
            }
        }

        logger.info(f"Created portfolio with {len(holdings)} holdings")
        return portfolio

    def save_portfolio(self, portfolio_data: Dict, output_path: str) -> None:
        """
        Save portfolio to JSON file

        Args:
            portfolio_data: Portfolio data dictionary
            output_path: Output JSON file path
        """
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(portfolio_data, f, indent=2)

        logger.info(f"Saved portfolio to {output_path}")


def import_from_csv(holdings_csv: str, transactions_csv: Optional[str] = None,
                   output_json: Optional[str] = None, pan: str = "IMPORTED") -> Dict:
    """
    Convenience function to import from CSV

    Args:
        holdings_csv: Path to holdings CSV file
        transactions_csv: Optional path to transactions CSV file
        output_json: Optional output JSON file path
        pan: PAN number for the account

    Returns:
        Portfolio data dictionary
    """
    importer = DataImporter()
    portfolio = importer.csv_to_portfolio(holdings_csv, transactions_csv, pan)

    if output_json:
        importer.save_portfolio(portfolio, output_json)

    return portfolio


def import_from_excel(excel_path: str, output_json: Optional[str] = None,
                     pan: str = "IMPORTED") -> Dict:
    """
    Convenience function to import from Excel

    Args:
        excel_path: Path to Excel file
        output_json: Optional output JSON file path
        pan: PAN number for the account

    Returns:
        Portfolio data dictionary
    """
    importer = DataImporter()
    portfolio = importer.excel_to_portfolio(excel_path, pan)

    if output_json:
        importer.save_portfolio(portfolio, output_json)

    return portfolio


if __name__ == "__main__":
    print("Data Import Module")
    print("=" * 50)
    print("\nExample usage:")
    print("""
from backend.data_import import DataImporter, import_from_csv, import_from_excel

# Import from CSV
portfolio = import_from_csv(
    holdings_csv='data/output/holdings.csv',
    transactions_csv='data/output/transactions.csv',
    output_json='data/imported_portfolio.json',
    pan='ABCDE1234F'
)

# Import from Excel
portfolio = import_from_excel(
    excel_path='data/output/portfolio.xlsx',
    output_json='data/imported_portfolio.json',
    pan='ABCDE1234F'
)

# Using DataImporter class
importer = DataImporter()

# Import holdings
holdings = importer.import_holdings_csv('holdings.csv')

# Import transactions
transactions = importer.import_transactions_csv('transactions.csv')

# Convert to portfolio format
portfolio = importer.csv_to_portfolio(
    holdings_csv='holdings.csv',
    transactions_csv='transactions.csv',
    pan='ABCDE1234F',
    account_name='My Account'
)

# Save portfolio
importer.save_portfolio(portfolio, 'my_portfolio.json')
""")

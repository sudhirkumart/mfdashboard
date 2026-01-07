"""
Excel Export Module
Exports portfolio data to Excel with formulas and formatting
"""

from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, BarChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("openpyxl not installed. Excel export will not be available.")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export portfolio data to Excel with formulas and formatting"""

    def __init__(self):
        """Initialize Excel exporter"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        self.workbook = None
        self.styles = self._create_styles()

    def _create_styles(self) -> Dict:
        """Create reusable cell styles"""
        return {
            'header': {
                'font': Font(bold=True, color='FFFFFF', size=11),
                'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'title': {
                'font': Font(bold=True, size=14),
                'alignment': Alignment(horizontal='center')
            },
            'currency': {
                'alignment': Alignment(horizontal='right')
            },
            'percentage': {
                'alignment': Alignment(horizontal='right')
            },
            'date': {
                'alignment': Alignment(horizontal='center')
            }
        }

    def _apply_header_style(self, cell) -> None:
        """Apply header style to a cell"""
        for key, value in self.styles['header'].items():
            setattr(cell, key, value)

    def export_portfolio(self, portfolio_data: Dict, output_path: str) -> None:
        """
        Export portfolio to Excel

        Args:
            portfolio_data: Portfolio data dictionary
            output_path: Output Excel file path
        """
        logger.info(f"Exporting portfolio to Excel: {output_path}")

        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet

        # Create sheets
        self._create_summary_sheet(portfolio_data)
        self._create_holdings_sheet(portfolio_data)
        self._create_transactions_sheet(portfolio_data)
        self._create_pan_wise_sheet(portfolio_data)

        # Save workbook
        self.workbook.save(output_path)
        logger.info(f"Portfolio exported successfully to {output_path}")

    def _create_summary_sheet(self, portfolio_data: Dict) -> None:
        """Create portfolio summary sheet"""
        ws = self.workbook.create_sheet("Summary", 0)

        # Title
        ws['A1'] = "Portfolio Summary"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        # Portfolio info
        row = 3
        ws[f'A{row}'] = "Portfolio Name:"
        ws[f'B{row}'] = portfolio_data.get('portfolio_name', 'My Portfolio')
        ws[f'B{row}'].font = Font(bold=True)

        row += 1
        ws[f'A{row}'] = "Generated Date:"
        ws[f'B{row}'] = datetime.now().strftime('%d-%b-%Y %H:%M')

        row += 2
        ws[f'A{row}'] = "Total Accounts:"
        ws[f'B{row}'] = len(portfolio_data.get('accounts', {}))

        # Summary metrics
        row += 2
        metrics_headers = ['Metric', 'Value']
        for col, header in enumerate(metrics_headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_header_style(cell)

        # Calculate totals
        total_invested = 0
        total_current = 0

        for pan, account in portfolio_data.get('accounts', {}).items():
            for holding in account.get('holdings', []):
                total_invested += holding.get('invested', 0)
                total_current += holding.get('current_value', 0)

        total_gain = total_current - total_invested
        return_pct = (total_gain / total_invested * 100) if total_invested > 0 else 0

        # Add metrics
        metrics = [
            ('Total Invested', total_invested, '₹#,##0.00'),
            ('Current Value', total_current, '₹#,##0.00'),
            ('Total Gain/Loss', total_gain, '₹#,##0.00'),
            ('Return %', return_pct, '0.00%')
        ]

        row += 1
        for metric_name, value, format_str in metrics:
            ws[f'A{row}'] = metric_name
            ws[f'B{row}'] = value
            if format_str:
                ws[f'B{row}'].number_format = format_str
            row += 1

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20

    def _create_holdings_sheet(self, portfolio_data: Dict) -> None:
        """Create detailed holdings sheet with formulas"""
        ws = self.workbook.create_sheet("Holdings")

        # Headers
        headers = [
            'PAN', 'Scheme Name', 'Folio', 'Units',
            'Avg Buy Price', 'Invested', 'Current NAV',
            'Current Value', 'Gain/Loss', 'Return %', 'NAV Date'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Add holdings data
        row = 2
        for pan, account in portfolio_data.get('accounts', {}).items():
            for holding in account.get('holdings', []):
                ws.cell(row=row, column=1, value=pan)
                ws.cell(row=row, column=2, value=holding.get('scheme_name', ''))
                ws.cell(row=row, column=3, value=holding.get('folio', ''))
                ws.cell(row=row, column=4, value=holding.get('units', 0))

                invested = holding.get('invested', 0)
                units = holding.get('units', 0)
                avg_price = invested / units if units > 0 else 0

                ws.cell(row=row, column=5, value=avg_price)
                ws.cell(row=row, column=6, value=invested)
                ws.cell(row=row, column=7, value=holding.get('current_nav', 0))

                # Current Value formula: Units * Current NAV
                current_val_col = get_column_letter(8)
                units_col = get_column_letter(4)
                nav_col = get_column_letter(7)
                ws[f'{current_val_col}{row}'] = f'={units_col}{row}*{nav_col}{row}'

                # Gain/Loss formula: Current Value - Invested
                gain_col = get_column_letter(9)
                invested_col = get_column_letter(6)
                ws[f'{gain_col}{row}'] = f'={current_val_col}{row}-{invested_col}{row}'

                # Return % formula: (Gain/Loss / Invested) * 100
                return_col = get_column_letter(10)
                ws[f'{return_col}{row}'] = f'=IF({invested_col}{row}>0,({gain_col}{row}/{invested_col}{row})*100,0)'

                ws.cell(row=row, column=11, value=holding.get('nav_date', ''))

                # Apply number formats
                ws.cell(row=row, column=5).number_format = '₹#,##0.00'
                ws.cell(row=row, column=6).number_format = '₹#,##0.00'
                ws.cell(row=row, column=7).number_format = '₹#,##0.00'
                ws.cell(row=row, column=8).number_format = '₹#,##0.00'
                ws.cell(row=row, column=9).number_format = '₹#,##0.00'
                ws.cell(row=row, column=10).number_format = '0.00%'

                row += 1

        # Add totals row
        if row > 2:
            ws[f'A{row}'] = 'TOTAL'
            ws[f'A{row}'].font = Font(bold=True)

            # Sum formulas
            ws[f'F{row}'] = f'=SUM(F2:F{row-1})'
            ws[f'H{row}'] = f'=SUM(H2:H{row-1})'
            ws[f'I{row}'] = f'=SUM(I2:I{row-1})'

            ws[f'F{row}'].number_format = '₹#,##0.00'
            ws[f'H{row}'].number_format = '₹#,##0.00'
            ws[f'I{row}'].number_format = '₹#,##0.00'

            for col in range(1, 12):
                ws.cell(row=row, column=col).font = Font(bold=True)

        # Auto-adjust column widths
        for col in range(1, 12):
            ws.column_dimensions[get_column_letter(col)].width = 15

        ws.column_dimensions['B'].width = 35  # Scheme name

    def _create_transactions_sheet(self, portfolio_data: Dict) -> None:
        """Create transactions sheet"""
        ws = self.workbook.create_sheet("Transactions")

        # Headers
        headers = ['Date', 'PAN', 'Scheme Name', 'Type', 'Units', 'NAV', 'Amount']

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Add transactions
        row = 2
        for pan, account in portfolio_data.get('accounts', {}).items():
            for holding in account.get('holdings', []):
                for txn in holding.get('transactions', []):
                    ws.cell(row=row, column=1, value=txn.get('date', ''))
                    ws.cell(row=row, column=2, value=pan)
                    ws.cell(row=row, column=3, value=holding.get('scheme_name', ''))
                    ws.cell(row=row, column=4, value=txn.get('type', 'purchase').upper())
                    ws.cell(row=row, column=5, value=txn.get('units', 0))
                    ws.cell(row=row, column=6, value=txn.get('nav', 0))

                    # Amount formula
                    ws[f'G{row}'] = f'=E{row}*F{row}'

                    # Apply formats
                    ws.cell(row=row, column=6).number_format = '₹#,##0.00'
                    ws.cell(row=row, column=7).number_format = '₹#,##0.00'

                    row += 1

        # Auto-adjust columns
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 12

    def _create_pan_wise_sheet(self, portfolio_data: Dict) -> None:
        """Create PAN-wise summary sheet"""
        ws = self.workbook.create_sheet("PAN Summary")

        # Headers
        headers = ['PAN', 'Name', 'Holdings Count', 'Invested', 'Current Value', 'Gain/Loss', 'Return %']

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Add PAN data
        row = 2
        for pan, account in portfolio_data.get('accounts', {}).items():
            invested = 0
            current_value = 0

            for holding in account.get('holdings', []):
                invested += holding.get('invested', 0)
                current_value += holding.get('current_value', 0)

            gain_loss = current_value - invested
            return_pct = (gain_loss / invested * 100) if invested > 0 else 0

            ws.cell(row=row, column=1, value=pan)
            ws.cell(row=row, column=2, value=account.get('name', ''))
            ws.cell(row=row, column=3, value=len(account.get('holdings', [])))
            ws.cell(row=row, column=4, value=invested)
            ws.cell(row=row, column=5, value=current_value)
            ws.cell(row=row, column=6, value=gain_loss)
            ws.cell(row=row, column=7, value=return_pct / 100)

            # Apply formats
            ws.cell(row=row, column=4).number_format = '₹#,##0.00'
            ws.cell(row=row, column=5).number_format = '₹#,##0.00'
            ws.cell(row=row, column=6).number_format = '₹#,##0.00'
            ws.cell(row=row, column=7).number_format = '0.00%'

            row += 1

        # Auto-adjust columns
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15

        ws.column_dimensions['B'].width = 25


def export_to_excel(portfolio_data: Dict, output_path: str) -> None:
    """
    Convenience function to export portfolio to Excel

    Args:
        portfolio_data: Portfolio data dictionary
        output_path: Output Excel file path
    """
    exporter = ExcelExporter()
    exporter.export_portfolio(portfolio_data, output_path)


if __name__ == "__main__":
    print("Excel Export Module")
    print("=" * 50)

    if not EXCEL_AVAILABLE:
        print("\nWARNING: openpyxl not installed!")
        print("Install with: pip install openpyxl")
        print("\nExample usage (when openpyxl is installed):")
    else:
        print("\nExample usage:")

    print("""
from backend.excel_export import ExcelExporter

# Initialize exporter
exporter = ExcelExporter()

# Export portfolio
exporter.export_portfolio(portfolio_data, 'output/portfolio.xlsx')

# Or use convenience function
from backend.excel_export import export_to_excel
export_to_excel(portfolio_data, 'output/portfolio.xlsx')
""")

    print("\nThe generated Excel file will contain:")
    print("- Summary sheet with portfolio overview")
    print("- Holdings sheet with detailed fund information and formulas")
    print("- Transactions sheet with all buy/sell records")
    print("- PAN Summary sheet with account-wise breakdown")
